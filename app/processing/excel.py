from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font


def build_excel_workbook(
    participants: List[Dict[str, Any]],
    mentions: List[Dict[str, Any]],
    channels: List[Dict[str, Any]],
    export_date: datetime,
) -> Workbook:
    """
    Формирует Excel с вкладками:
    - Участники
    - Упоминания
    - Каналы

    Поля (для Участников):
    - Дата экспорта
    - Username
    - Имя и фамилия
    - Описание (Bio/About)
    - Дата регистрации
    - Наличие канала в профиле
    """
    wb = Workbook()

    # Участники
    ws_part = wb.active
    ws_part.title = "Участники"
    headers = ["Дата экспорта", "Username", "Имя и фамилия", "Описание", "Дата регистрации", "Наличие канала в профиле"]
    ws_part.append(headers)
    for cell in ws_part[1]:
        cell.font = Font(bold=True)

    for p in participants:
        ws_part.append(
            [
                export_date.strftime("%Y-%m-%d"),
                (p.get("username") or "") if p.get("username") else "",
                (p.get("name") or "") if p.get("name") else "",
                (p.get("bio") or "") if p.get("bio") else "",
                (p.get("registered_at") or "") if p.get("registered_at") else "",
                "Да" if p.get("has_channel") else "Нет",
            ]
        )

    # Упоминания
    ws_mentions = wb.create_sheet(title="Упоминания")
    ws_mentions.append(["Username"])
    ws_mentions["A1"].font = Font(bold=True)
    for m in mentions:
        uname = m.get("username")
        if uname:
            ws_mentions.append([f"@{uname}" if not uname.startswith("@") else uname])

    # Каналы
    ws_channels = wb.create_sheet(title="Каналы")
    ws_channels.append(["Name", "Username"])
    ws_channels["A1"].font = Font(bold=True)
    ws_channels["B1"].font = Font(bold=True)
    for c in channels:
        name = c.get("name") or ""
        uname = c.get("username") or ""
        ws_channels.append([name, f"@{uname}" if uname and not uname.startswith("@") else uname])

    return wb